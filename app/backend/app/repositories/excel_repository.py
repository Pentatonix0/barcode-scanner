import asyncio
from dataclasses import asdict
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any

from app.core.core_utils.models import Product
from app.repositories.base import ProductRepository
from openpyxl import load_workbook


class ExcelProductRepository(ProductRepository):
    def __init__(self, file_path: Path, barcode_column: str = "barcode") -> None:
        self._file_path = Path(file_path)
        self._barcode_column = barcode_column.lower().strip()
        self._items: dict[str, Product] = {}
        self._columns: list[str] = []
        self._barcode_header: str | None = None
        self._last_loaded_at: datetime | None = None
        self._lock = asyncio.Lock()

    async def get_by_barcode(self, barcode: str) -> Product | None:
        normalized = _normalize_barcode(barcode)
        if not normalized:
            return None
        return self._items.get(normalized)

    async def reload(self) -> None:
        if not self._file_path.exists():
            self._items = {}
            self._columns = []
            self._barcode_header = None
            self._last_loaded_at = None
            return

        async with self._lock:
            items, columns, barcode_header = _load_catalog(
                self._file_path, self._barcode_column
            )
            self._items = items
            self._columns = columns
            self._barcode_header = barcode_header
            self._last_loaded_at = datetime.now()

    def meta(self) -> dict:
        return {
            "count": len(self._items),
            "columns": self._columns,
            "barcode_column": self._barcode_header or self._barcode_column,
            "last_loaded_at": (
                self._last_loaded_at.isoformat() if self._last_loaded_at else None
            ),
            "file": str(self._file_path),
        }


def _load_catalog(
    file_path: Path, barcode_column: str
) -> tuple[dict[str, Product], list[str], str]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active

    rows = sheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return {}, [], ""

    normalized_headers = _normalize_headers(headers)
    barcode_index = _find_barcode_index(normalized_headers, barcode_column)
    if barcode_index is None:
        raise ValueError(f"Barcode column '{barcode_column}' not found")

    items: dict[str, Product] = {}

    for row in rows:
        if row is None:
            continue
        if all(cell is None for cell in row):
            continue

        row_values = _row_to_dict(normalized_headers, row)
        barcode_value = _normalize_barcode(row[barcode_index])
        if not barcode_value:
            continue

        fields = {
            key: value
            for key, value in row_values.items()
            if key != normalized_headers[barcode_index]
        }
        items[barcode_value] = Product(barcode=barcode_value, fields=fields)

    columns = [name for name in normalized_headers if name]
    return items, columns, normalized_headers[barcode_index]


def _normalize_headers(headers: tuple[Any, ...]) -> list[str]:
    seen: dict[str, int] = {}
    normalized: list[str] = []

    for header in headers:
        name = str(header).strip() if header is not None else ""
        if not name:
            normalized.append("")
            continue

        count = seen.get(name, 0) + 1
        seen[name] = count
        if count > 1:
            name = f"{name}_{count}"
        normalized.append(name)

    return normalized


def _find_barcode_index(headers: list[str], barcode_column: str) -> int | None:
    for idx, name in enumerate(headers):
        if name.lower() == barcode_column:
            return idx
    return None


def _row_to_dict(headers: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for idx, name in enumerate(headers):
        if not name:
            continue
        value = row[idx] if idx < len(row) else None
        result[name] = _normalize_cell_value(value)
    return result


def _normalize_barcode(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).rstrip("0").rstrip(".")
    if isinstance(value, (int, Decimal)):
        return str(value)
    text = str(value).strip()
    if text.endswith(".0") and text.replace(".0", "").isdigit():
        return text[:-2]
    return text


def _normalize_cell_value(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, Product):
        return asdict(value)
    return value
